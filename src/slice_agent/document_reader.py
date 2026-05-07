"""Document reading utilities for various file formats."""

from pathlib import Path
from typing import Dict, Any


def read_document(file_path: str) -> Dict[str, Any]:
    """
    Read and extract text content from various document formats.

    Supported formats:
    - PDF (.pdf)
    - Word (.docx)
    - Excel (.xlsx only - .xls not supported)
    - CSV (.csv)
    - Plain text (.txt, .md, .py, .js, .json, etc.)

    Args:
        file_path: Path to the document to read

    Returns:
        Dict with 'success', 'content', 'error', 'file_type' keys
    """
    path = Path(file_path)

    # Check if file exists
    if not path.exists():
        return {
            "success": False,
            "content": "",
            "error": f"File not found: {file_path}",
            "file_type": "unknown"
        }

    # Check if it's a file (not a directory)
    if not path.is_file():
        return {
            "success": False,
            "content": "",
            "error": f"Not a file: {file_path}",
            "file_type": "unknown"
        }

    # Determine file type and read accordingly
    suffix = path.suffix.lower()

    try:
        if suffix == '.pdf':
            content = _read_pdf(path)
            file_type = "PDF"
        elif suffix == '.docx':
            content = _read_docx(path)
            file_type = "Word Document"
        elif suffix == '.xlsx':
            content = _read_excel(path)
            file_type = "Excel Spreadsheet"
        elif suffix == '.xls':
            # Old Excel format - not supported by openpyxl
            raise ValueError(
                "Legacy .xls format is not supported. "
                "Please convert to .xlsx or use a different tool."
            )
        elif suffix == '.csv':
            content = _read_csv(path)
            file_type = "CSV File"
        else:
            # Try to read as plain text
            content = _read_text(path)
            file_type = "Text Document"

        return {
            "success": True,
            "content": content,
            "error": "",
            "file_type": file_type
        }
    except Exception as e:
        return {
            "success": False,
            "content": "",
            "error": f"Failed to read {file_path}: {str(e)}",
            "file_type": suffix[1:] if suffix else "unknown"
        }


def _read_pdf(path: Path) -> str:
    """Read text content from a PDF file."""
    try:
        from pypdf import PdfReader
    except ImportError:
        raise ImportError(
            "pypdf is required to read PDF files. "
            "Please reinstall slice-agent: pip install -e ."
        )

    reader = PdfReader(path)
    text_content = []

    for page_num, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text()
        if page_text.strip():
            text_content.append(f"--- Page {page_num} ---\n{page_text}")

    if not text_content:
        return "(No text content could be extracted from this PDF)"

    return "\n\n".join(text_content)


def _read_docx(path: Path) -> str:
    """Read text content from a Word document."""
    try:
        from docx import Document
    except ImportError:
        raise ImportError(
            "python-docx is required to read Word documents. "
            "Please reinstall slice-agent: pip install -e ."
        )

    doc = Document(path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]

    if not paragraphs:
        return "(No text content could be extracted from this Word document)"

    return "\n\n".join(paragraphs)


def _read_excel(path: Path) -> str:
    """Read text content from an Excel file (.xlsx only)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required to read Excel files. "
            "Please reinstall slice-agent: pip install -e ."
        )

    # data_only=True means formulas are evaluated to their values
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheets_content = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Get dimensions to avoid reading empty cells
        if sheet.max_row == 0 or sheet.max_column == 0:
            continue

        sheet_data = []
        sheet_data.append(f"=== Sheet: {sheet_name} ===\n")

        # Read all rows
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Convert row values to strings, handling None
            row_values = [str(cell) if cell is not None else "" for cell in row]
            # Only include rows that have at least one non-empty cell
            if any(val.strip() for val in row_values):
                # Format as "Row N: value1 | value2 | value3"
                row_text = f"Row {row_num}: " + " | ".join(row_values)
                sheet_data.append(row_text)

        if len(sheet_data) > 1:  # More than just the header
            sheets_content.append("\n".join(sheet_data))

    if not sheets_content:
        return "(No data could be extracted from this Excel file)"

    return "\n\n".join(sheets_content)


def _read_csv(path: Path) -> str:
    """Read a CSV file."""
    import csv

    rows_content = []
    last_error = None

    # Try UTF-8 first, then latin-1
    for encoding in ['utf-8', 'latin-1']:
        try:
            with open(path, 'r', encoding=encoding, newline='') as f:
                reader = csv.reader(f)
                for row_num, row in enumerate(reader, start=1):
                    if any(cell.strip() for cell in row):  # Skip empty rows
                        row_text = f"Row {row_num}: " + " | ".join(row)
                        rows_content.append(row_text)
            break  # Success, exit the encoding loop
        except UnicodeDecodeError as e:
            last_error = e
            continue  # Try next encoding
        except Exception as e:
            # Other errors (permissions, CSV format issues, etc.)
            raise

    if not rows_content:
        if last_error:
            raise last_error
        return "(No data could be extracted from this CSV file)"

    return "\n".join(rows_content)


def _read_text(path: Path) -> str:
    """Read a plain text file with automatic encoding detection."""
    # Try UTF-8 first (most common)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return f.read()
    except UnicodeDecodeError:
        # Fall back to latin-1 which accepts all byte values
        try:
            with open(path, 'r', encoding='latin-1') as f:
                return f.read()
        except Exception:
            # Last resort: read as binary and decode with errors='replace'
            with open(path, 'rb') as f:
                return f.read().decode('utf-8', errors='replace')
